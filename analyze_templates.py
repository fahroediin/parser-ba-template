#!/usr/bin/env python3

import pandas as pd
import openpyxl
from io import BytesIO

def analyze_template_structure(file_path, template_name):
    """Analyze Excel template structure"""

    print(f"\n{'='*60}")
    print(f"ANALYSIS: {template_name}")
    print(f"File: {file_path}")
    print(f"{'='*60}")

    try:
        # Load Excel file
        workbook = openpyxl.load_workbook(file_path)

        print(f"\n📋 Available Sheets:")
        for i, sheet_name in enumerate(workbook.sheet_names, 1):
            worksheet = workbook[sheet_name]
            print(f"  {i}. {sheet_name} (Used Range: {worksheet.calculate_dimension()})")

        print(f"\n🔍 Detailed Sheet Analysis:")

        for sheet_name in workbook.sheet_names:
            print(f"\n📄 Sheet: {sheet_name}")
            print("-" * 40)

            worksheet = workbook[sheet_name]

            # Check for images in this sheet
            if hasattr(worksheet, '_images') and worksheet._images:
                print(f"  🖼️  Images Found: {len(worksheet._images)}")
                for i, img in enumerate(worksheet._images[:3], 1):  # Show first 3
                    print(f"    Image {i}: {type(img).__name__}")
                    if hasattr(img, 'anchor'):
                        print(f"      Location: {img.anchor}")

            # Read first few rows to understand structure
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=10)
                print(f"  📊 Sample Data (First 10 rows):")

                for row_idx, row in df.iterrows():
                    row_data = []
                    for col_idx, cell in enumerate(row):
                        if pd.notna(cell):
                            row_data.append(f"Col{col_idx+1}:{str(cell)[:30]}")
                        else:
                            row_data.append("Col{col_idx+1}:empty")

                    if row_data:  # Only show rows with data
                        print(f"    Row {row_idx+1}: {' | '.join(row_data[:3])}...")

            except Exception as e:
                print(f"  ❌ Error reading sheet data: {e}")

            # Check for merged cells
            try:
                if worksheet.merged_cells.ranges:
                    print(f"  🔀 Merged Cells: {len(worksheet.merged_cells.ranges)}")
                    for i, merged_range in enumerate(worksheet.merged_cells.ranges[:3], 1):
                        print(f"    {i}: {merged_range}")
            except:
                pass

    except Exception as e:
        print(f"❌ Error analyzing template: {e}")
        return False

    return True

def analyze_uix_template():
    """Analyze UIUX template specifically"""

    print(f"\n{'='*80}")
    print(f"UIUX TEMPLATE ANALYSIS")
    print(f"{'='*80}")

    uix_path = r"C:\Users\User\Downloads\Template_UIUX_CatalogApp.xlsx"

    # Load with pandas for quick overview
    try:
        uix_workbook = openpyxl.load_workbook(uix_path)

        print(f"\n🎨 UIUX Template Features:")

        # Check each sheet for UIUX-specific content
        for sheet_name in uix_workbook.sheet_names:
            worksheet = uix_workbook[sheet_name]
            print(f"\n📋 {sheet_name}:")

            # Load with pandas for data analysis
            try:
                df = pd.read_excel(uix_path, sheet_name=sheet_name, header=None)

                # Look for UIUX-specific keywords
                uix_keywords = ['design', 'mockup', 'prototype', 'wireframe', 'figma', 'sketch', 'layout', 'component', 'user interface']
                found_keywords = []

                for _, row in df.iterrows():
                    for cell in row:
                        if pd.notna(cell):
                            cell_str = str(cell).lower()
                            for keyword in uix_keywords:
                                if keyword in cell_str and keyword not in found_keywords:
                                    found_keywords.append(keyword)

                if found_keywords:
                    print(f"  🎨 UIUX Keywords Found: {', '.join(found_keywords)}")

                # Check for design file references
                file_extensions = ['.png', '.jpg', '.jpeg', '.gif', '.svg', '.figma', '.sketch', '.psd', '.ai']
                found_files = []

                for _, row in df.iterrows():
                    for cell in row:
                        if pd.notna(cell):
                            cell_str = str(cell)
                            for ext in file_extensions:
                                if ext in cell_str and ext not in [f.split(ext)[-1] for f in found_files]:
                                    found_files.append(cell_str)

                if found_files:
                    print(f"  📁 Design Files Referenced: {found_files[:3]}")  # Show first 3

                # Look for image metadata
                image_keywords = ['image', 'screenshot', 'attachment', 'asset', 'mockup', 'design', 'prototype']
                image_rows = []

                for idx, row in df.iterrows():
                    row_str = ' '.join([str(cell) for cell in row if pd.notna(cell)]).lower()
                    if any(keyword in row_str for keyword in image_keywords):
                        image_rows.append((idx+1, row_str))

                if image_rows:
                    print(f"  🖼️  Image Metadata Rows: {len(image_rows)}")
                    for row_num, row_data in image_rows[:2]:  # Show first 2
                        print(f"    Row {row_num}: {row_data[:80]}...")

            except Exception as e:
                print(f"  ❌ Error analyzing {sheet_name}: {e}")

    except Exception as e:
        print(f"❌ Error loading UIUX template: {e}")

def analyze_engineer_template():
    """Analyze Engineer template specifically"""

    print(f"\n{'='*80}")
    print(f"ENGINEER TEMPLATE ANALYSIS")
    print(f"{'='*80}")

    eng_path = r"C:\Users\User\Downloads\Template_Dev_CatalogApp.xlsx"

    try:
        eng_workbook = openpyxl.load_workbook(eng_path)

        print(f"\n⚙️ Engineer Template Features:")

        for sheet_name in eng_workbook.sheet_names:
            worksheet = eng_workbook[sheet_name]
            print(f"\n📋 {sheet_name}:")

            try:
                df = pd.read_excel(eng_path, sheet_name=sheet_name, header=None)

                # Look for engineering-specific keywords
                eng_keywords = ['api', 'database', 'architecture', 'tech stack', 'infrastructure', 'deployment', 'server', 'database', 'backend', 'frontend', 'devops']
                found_keywords = []

                for _, row in df.iterrows():
                    for cell in row:
                        if pd.notna(cell):
                            cell_str = str(cell).lower()
                            for keyword in eng_keywords:
                                if keyword in cell_str and keyword not in found_keywords:
                                    found_keywords.append(keyword)

                if found_keywords:
                    print(f"  ⚙️ Engineering Keywords Found: {', '.join(found_keywords)}")

                # Check for technical diagrams
                diagram_keywords = ['diagram', 'architecture', 'flowchart', 'schema', 'erd', 'network', 'system']
                diagram_rows = []

                for idx, row in df.iterrows():
                    row_str = ' '.join([str(cell) for cell in row if pd.notna(cell)]).lower()
                    if any(keyword in row_str for keyword in diagram_keywords):
                        diagram_rows.append((idx+1, row_str))

                if diagram_rows:
                    print(f"  📐 Technical Diagram References: {len(diagram_rows)}")
                    for row_num, row_data in diagram_rows[:2]:
                        print(f"    Row {row_num}: {row_data[:80]}...")

            except Exception as e:
                print(f"  ❌ Error analyzing {sheet_name}: {e}")

    except Exception as e:
        print(f"❌ Error loading Engineer template: {e}")

def create_modular_design():
    """Design modular parser architecture"""

    print(f"\n{'='*80}")
    print(f"MODULAR PARSER DESIGN")
    print(f"{'='*80}")

    print(f"""
🏗️ RECOMMENDED ARCHITECTURE:

1. 📁 Core Parser Structure:
   /services/
   ├── base_parser.py           # Base class with common functionality
   ├── ba_parser.py            # Business Analysis parser (current)
   ├── uix_parser.py           # UI/UX specific parser
   ├── eng_parser.py           # Engineering specific parser
   ├── image_extractor.py      # Image extraction (shared)
   └── parser_factory.py       # Factory to instantiate correct parser

2. 🎯 Division-Specific Features:

   BA Parser:
   - Product Overview, User Stories, AC, Business Value
   - BA Approval workflow
   - Text-focused parsing

   UIUX Parser:
   - Design Overview, Figma Links, Design Assets
   - Mockup image extraction
   - Component specifications
   - User journey flows

   Engineering Parser:
   - Tech Stack, Architecture Diagrams
   - Database Schema, API Endpoints
   - Infrastructure diagrams
   - Deployment specifications

3. 🔧 Parser Detection:
   - Auto-detect template type based on sheet names
   - Use factory pattern for parser instantiation
   - Fallback to BA parser if unknown

4. 📊 Unified API:
   - Single /upload endpoint
   - Auto-select parser based on file content
   - Unified response format across divisions
""")

if __name__ == "__main__":
    # Analyze templates
    uix_success = analyze_template_structure(r"C:\Users\User\Downloads\Template_UIUX_CatalogApp.xlsx", "UIUX Template")
    eng_success = analyze_template_structure(r"C:\Users\User\Downloads\Template_Dev_CatalogApp.xlsx", "Engineer Template")

    analyze_uix_template()
    analyze_engineer_template()
    create_modular_design()

    if uix_success and eng_success:
        print(f"\n✅ TEMPLATE ANALYSIS COMPLETE!")
        print(f"🚀 Ready to implement modular parser architecture")
    else:
        print(f"\n❌ TEMPLATE ANALYSIS FAILED!")