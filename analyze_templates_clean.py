#!/usr/bin/env python3

import pandas as pd
import openpyxl
import os

def analyze_template_file(file_path, template_name):
    """Clean template analysis without Unicode characters"""

    print(f"\n{'='*60}")
    print(f"ANALYSIS: {template_name}")
    print(f"File: {file_path}")
    print(f"{'='*60}")

    if not os.path.exists(file_path):
        print(f"ERROR: File not found at {file_path}")
        return False

    try:
        # Load with openpyxl for sheet names and structure
        workbook = openpyxl.load_workbook(file_path)

        print(f"\nSheets Found: {len(workbook.sheetnames)}")
        for i, sheet_name in enumerate(workbook.sheetnames, 1):
            worksheet = workbook[sheet_name]
            try:
                dimension = worksheet.calculate_dimension()
                print(f"  {i}. {sheet_name} (Range: {dimension})")
            except:
                print(f"  {i}. {sheet_name} (Empty or error)")

        # Detailed analysis of each sheet
        print(f"\nDetailed Sheet Analysis:")
        for sheet_name in workbook.sheetnames:
            print(f"\nSheet: {sheet_name}")
            print("-" * 40)

            worksheet = workbook[sheet_name]

            # Check for images
            if hasattr(worksheet, '_images') and worksheet._images:
                print(f"  Images: {len(worksheet._images)} embedded")
            else:
                print(f"  Images: None detected")

            # Load sample data with pandas
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=10)
                print(f"  Data shape: {df.shape}")

                # Show first few non-empty rows
                non_empty_rows = 0
                for idx, row in df.iterrows():
                    if any(pd.notna(cell) for cell in row):
                        if non_empty_rows < 3:  # Show first 3 rows with data
                            row_preview = []
                            for cell in row[:3]:  # First 3 columns
                                if pd.notna(cell):
                                    row_preview.append(str(cell)[:30])
                            if row_preview:
                                print(f"    Row {idx+1}: {' | '.join(row_preview)}")
                        non_empty_rows += 1

                print(f"  Non-empty rows: {non_empty_rows}")

            except Exception as e:
                print(f"  Error reading data: {e}")

        return True

    except Exception as e:
        print(f"ERROR analyzing template: {e}")
        return False

def detect_template_patterns(file_path, template_name):
    """Detect specific patterns in templates"""

    print(f"\n{'='*60}")
    print(f"PATTERN DETECTION: {template_name}")
    print(f"{'='*60}")

    try:
        # Check specific division patterns
        uix_patterns = ['design', 'figma', 'mockup', 'wireframe', 'prototype', 'ui', 'ux', 'component', 'layout']
        eng_patterns = ['tech', 'stack', 'api', 'database', 'architecture', 'infrastructure', 'deployment', 'server', 'backend']
        ba_patterns = ['product', 'user story', 'acceptance criteria', 'business value', 'approval', 'requirement']

        workbook = openpyxl.load_workbook(file_path)
        sheet_names = workbook.sheetnames

        print(f"\nSheet Name Analysis:")
        for sheet_name in sheet_names:
            sheet_lower = sheet_name.lower()

            uix_matches = [p for p in uix_patterns if p in sheet_lower]
            eng_matches = [p for p in eng_patterns if p in sheet_lower]
            ba_matches = [p for p in ba_patterns if p in sheet_lower]

            print(f"  {sheet_name}:")
            if uix_matches:
                print(f"    UIUX patterns: {', '.join(uix_matches)}")
            if eng_matches:
                print(f"    Engineering patterns: {', '.join(eng_matches)}")
            if ba_matches:
                print(f"    BA patterns: {', '.join(ba_matches)}")
            if not (uix_matches or eng_matches or ba_matches):
                print(f"    No specific patterns detected")

        print(f"\nContent Analysis:")
        for sheet_name in sheet_names:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                content_text = ' '.join([str(cell) for cell in df.values.flatten() if pd.notna(cell)]).lower()

                uix_content = [p for p in uix_patterns if p in content_text]
                eng_content = [p for p in eng_patterns if p in content_text]
                ba_content = [p for p in ba_patterns if p in content_text]

                if uix_content or eng_content or ba_content:
                    print(f"  {sheet_name}:")
                    if uix_content:
                        print(f"    UIUX content: {', '.join(set(uix_content))}")
                    if eng_content:
                        print(f"    Engineering content: {', '.join(set(eng_content))}")
                    if ba_content:
                        print(f"    BA content: {', '.join(set(ba_content))}")

            except Exception as e:
                print(f"  Error analyzing {sheet_name}: {e}")

    except Exception as e:
        print(f"ERROR in pattern detection: {e}")

if __name__ == "__main__":
    print("TEMPLATE ANALYSIS FOR MODULAR PARSER DESIGN")
    print("=" * 80)

    # Template file paths
    uix_path = r"C:\Users\User\Downloads\Template_UIUX_CatalogApp.xlsx"
    eng_path = r"C:\Users\User\Downloads\Template_Dev_CatalogApp.xlsx"

    # Analyze UIUX template
    uix_success = analyze_template_file(uix_path, "UIUX Template")
    if uix_success:
        detect_template_patterns(uix_path, "UIUX Template")

    # Analyze Engineer template
    eng_success = analyze_template_file(eng_path, "Engineer Template")
    if eng_success:
        detect_template_patterns(eng_path, "Engineer Template")

    print(f"\n{'='*80}")
    print("ANALYSIS COMPLETE")
    if uix_success and eng_success:
        print("Both templates analyzed successfully!")
        print("Ready to design modular parser architecture.")
    else:
        print("Some templates failed to analyze. Check file paths and formats.")
    print(f"{'='*80}")