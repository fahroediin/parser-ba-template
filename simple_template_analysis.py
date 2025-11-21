#!/usr/bin/env python3

import pandas as pd
import openpyxl

def analyze_template_sheets(file_path, template_name):
    """Analyze template sheets without emoji"""

    print(f"\n{'='*60}")
    print(f"ANALYSIS: {template_name}")
    print(f"File: {file_path}")
    print(f"{'='*60}")

    try:
        workbook = openpyxl.load_workbook(file_path)

        print(f"\nAvailable Sheets:")
        for i, sheet_name in enumerate(workbook.sheet_names, 1):
            worksheet = workbook[sheet_name]
            try:
                dim = worksheet.calculate_dimension()
                print(f"  {i}. {sheet_name} (Range: {dim})")
            except:
                print(f"  {i}. {sheet_name} (Empty)")

        return workbook.sheet_names

    except Exception as e:
        print(f"Error analyzing template: {e}")
        return []

def analyze_uiux_template():
    """Analyze UIUX template specific sheets"""

    print(f"\n{'='*80}")
    print(f"UIUX TEMPLATE SHEET ANALYSIS")
    print(f"{'='*80}")

    uix_path = r"C:\Users\User\Downloads\Template_UIUX_CatalogApp.xlsx"

    try:
        workbook = openpyxl.load_workbook(uix_path)
        sheets = workbook.sheet_names

        print(f"\nUIUX Sheets Found: {len(sheets)}")

        for sheet_name in sheets:
            print(f"\nSheet: {sheet_name}")

            worksheet = workbook[sheet_name]

            # Check for images
            if hasattr(worksheet, '_images') and worksheet._images:
                print(f"  Images: {len(worksheet._images)} embedded")
            else:
                print(f"  Images: None detected")

            # Read sample data
            try:
                df = pd.read_excel(uix_path, sheet_name=sheet_name, header=None, nrows=10)
                print(f"  Data shape: {df.shape}")

                # Look for UIUX keywords in first 5 rows
                uix_keywords = ['design', 'mockup', 'figma', 'wireframe', 'layout', 'component', 'ui', 'ux', 'prototype']
                found_keywords = []

                for idx, row in df.iterrows():
                    for cell in row:
                        if pd.notna(cell):
                            cell_str = str(cell).lower()
                            for keyword in uix_keywords:
                                if keyword in cell_str and keyword not in found_keywords:
                                    found_keywords.append(keyword)
                                    break
                    if found_keywords:  # Stop after finding some keywords
                        break

                if found_keywords:
                    print(f"  UIUX Keywords: {', '.join(found_keywords[:5])}")

                # Look for file references
                file_patterns = ['.png', '.jpg', '.jpeg', '.figma', '.sketch', '.psd']
                found_files = []

                for idx, row in df.iterrows():
                    for cell in row:
                        if pd.notna(cell):
                            cell_str = str(cell).lower()
                            for pattern in file_patterns:
                                if pattern in cell_str and pattern not in found_files:
                                    found_files.append(cell_str)
                                    break
                    if found_files:
                        break

                if found_files:
                    print(f"  Design Files: {found_files[:3]}")

            except Exception as e:
                print(f"  Error reading sheet data: {e}")

    except Exception as e:
        print(f"Error loading UIUX template: {e}")

def analyze_engineer_template():
    """Analyze Engineer template specific sheets"""

    print(f"\n{'='*80}")
    print(f"ENGINEER TEMPLATE SHEET ANALYSIS")
    print(f"{'='*80}")

    eng_path = r"C:\Users\User\Downloads\Template_Dev_CatalogApp.xlsx"

    try:
        workbook = openpyxl.load_workbook(eng_path)
        sheets = workbook.sheet_names

        print(f"\nEngineer Sheets Found: {len(sheets)}")

        for sheet_name in sheets:
            print(f"\nSheet: {sheet_name}")

            worksheet = workbook[sheet_name]

            # Check for images
            if hasattr(worksheet, '_images') and worksheet._images:
                print(f"  Images: {len(worksheet._images)} embedded")
            else:
                print(f"  Images: None detected")

            # Read sample data
            try:
                df = pd.read_excel(eng_path, sheet_name=sheet_name, header=None, nrows=10)
                print(f"  Data shape: {df.shape}")

                # Look for engineering keywords
                eng_keywords = ['tech', 'stack', 'database', 'api', 'architecture', 'infrastructure', 'deployment', 'server', 'backend', 'devops']
                found_keywords = []

                for idx, row in df.iterrows():
                    for cell in row:
                        if pd.notna(cell):
                            cell_str = str(cell).lower()
                            for keyword in eng_keywords:
                                if keyword in cell_str and keyword not in found_keywords:
                                    found_keywords.append(keyword)
                                    break
                    if found_keywords:
                        break

                if found_keywords:
                    print(f"  Engineering Keywords: {', '.join(found_keywords[:5])}")

            except Exception as e:
                print(f"  Error reading sheet data: {e}")

    except Exception as e:
        print(f"Error loading Engineer template: {e}")

def suggest_modular_architecture():
    """Suggest modular architecture design"""

    print(f"\n{'='*80}")
    print(f"MODULAR PARSER ARCHITECTURE RECOMMENDATION")
    print(f"{'='*80}")

    print(f"""
RECOMMENDED STRUCTURE:

/services/
├── base_parser.py           # Common functionality
│   - Excel file loading
│   - Basic data cleaning
│   - Image extraction (shared)
│   - Error handling
│
├── ba_parser.py            # Business Analysis
│   - Product Overview parsing
│   - User Stories, AC parsing
│   - Business Value parsing
│   - BA Approval parsing
│
├── uix_parser.py           # UI/UX Design
│   - Design Overview parsing
│   - Figma Links parsing
│   - Design Assets handling
│   - Mockup image extraction
│   - Component specifications
│
├── eng_parser.py           # Engineering
│   - Tech Stack parsing
│   - Architecture diagrams
│   - Database Schema parsing
│   - API Specifications
│   - Infrastructure parsing
│
├── parser_factory.py       # Factory Pattern
│   - Auto-detect template type
│   - Instantiate correct parser
│   - Fallback handling
│
└── image_extractor.py      # (already implemented)

PARSER DETECTION STRATEGY:
1. Check sheet names for patterns:
   - BA Template: "Product Overview", "User Story", "Acceptance Criteria", "Business Value"
   - UIUX Template: "Design Overview", "Figma Links", "Design Assets"
   - Engineer Template: "Tech Stack", "Architecture", "Database Schema", "API Spec"

2. Check for division-specific content:
   - UIUX: Figma links, design assets, mockups
   - Engineer: Technical specifications, architecture diagrams
   - BA: Business value, approval workflows

3. Fallback to BA parser for unknown templates

IMAGE HANDLING STRATEGY:
1. BA: Text-focused (existing implementation)
2. UIUX: Design assets + Figma integration + image extraction
3. Engineer: Technical diagrams + architecture images

API ENDPOINTS:
- Single /upload endpoint
- Auto-detect template type
- Unified response format
- Division-specific metadata
""")

if __name__ == "__main__":
    print("Starting Template Analysis...")

    # Analyze both templates
    uix_sheets = analyze_template_sheets(r"C:\Users\User\Downloads\Template_UIUX_CatalogApp.xlsx", "UIUX Template")
    eng_sheets = analyze_template_sheets(r"C:\Users\User\Downloads\Template_Dev_CatalogApp.xlsx", "Engineer Template")

    analyze_uix_template()
    analyze_engineer_template()
    suggest_modular_architecture()

    print(f"\n{'='*80}")
    print(f"ANALYSIS COMPLETE!")
    print(f"{'='*80}")
    print(f"Ready to implement modular parser architecture.")